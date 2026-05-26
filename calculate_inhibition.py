"""
Calculate Percent Inhibition from RawInput.xlsx

Calculates % inhibition for each replicate using the formula:
    % Inhibition = (1 - (cell_value / control_value)) * 100

Control value is taken from the same replicate column.
Output includes % inhibition values and average column.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def calculate_percent_inhibition(
    input_path: str,
    output_path: str = "PercInhibition.xlsx"
):
    """
    Calculate % inhibition for each replicate.
    
    Formula: (1 - (cell_value / control_value)) * 100
    Control value is taken from the same replicate column.
    
    Args:
        input_path: Path to RawInput.xlsx
        output_path: Output file path
        
    Output structure:
        | Single drugs | Drugs | Rep1_% | Rep2_% | ... | Avg_% |
    """
    
    # Load input workbook
    wb_input = load_workbook(input_path)
    
    # Create output workbook
    wb_output = Workbook()
    wb_output.remove(wb_output.active)
    
    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    control_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    avg_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    for sheet_name in wb_input.sheetnames:
        ws_input = wb_input[sheet_name]
        
        # Truncate sheet name if needed (Excel limit 31 chars)
        output_sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
        ws_output = wb_output.create_sheet(title=output_sheet_name)
        
        print(f"Processing: {sheet_name}")
        
        # Read header row
        header_row = next(ws_input.iter_rows(min_row=1, max_row=1, values_only=True))
        
        # Identify replicate columns (starting from column 3)
        num_replicates = len(header_row) - 2
        
        # Create output header
        output_headers = ["Single drugs", "Drugs"] + [f"Rep{i+1}_%" for i in range(num_replicates)] + ["Avg_%"]
        
        # Write header
        for col_idx, header in enumerate(output_headers, start=1):
            cell = ws_output.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # Read all data rows
        data_rows = list(ws_input.iter_rows(min_row=2, values_only=True))
        
        # Extract control values (first data row)
        control_row = data_rows[0]
        control_values = list(control_row[2:2+num_replicates])
        
        # Process each row
        output_row_idx = 2
        
        for row_idx, row in enumerate(data_rows):
            col_a = row[0]
            col_b = row[1]
            replicate_values = list(row[2:2+num_replicates])
            
            # Calculate % inhibition for each replicate
            perc_inhibition_values = []
            
            for i, val in enumerate(replicate_values):
                control_val = control_values[i]
                
                # Handle edge cases
                if val is None or control_val is None:
                    perc_val = None
                elif control_val == 0:
                    perc_val = 0.0
                else:
                    perc_val = (1 - (val / control_val)) * 100
                    # Round to 2 decimal places
                    perc_val = round(perc_val, 2)
                
                perc_inhibition_values.append(perc_val)
            
            # Calculate average (excluding None values)
            valid_values = [v for v in perc_inhibition_values if v is not None]
            avg_perc = round(sum(valid_values) / len(valid_values), 2) if valid_values else None
            
            # Write to output
            cell_a = ws_output.cell(row=output_row_idx, column=1, value=col_a)
            cell_a.border = thin_border
            cell_a.alignment = center_align
            
            cell_b = ws_output.cell(row=output_row_idx, column=2, value=col_b)
            cell_b.border = thin_border
            cell_b.alignment = center_align
            
            # Apply control styling
            if col_a == "Control":
                cell_a.fill = control_fill
                cell_b.fill = control_fill
            
            # Write % inhibition values
            for i, perc_val in enumerate(perc_inhibition_values):
                cell = ws_output.cell(row=output_row_idx, column=3+i, value=perc_val)
                cell.border = thin_border
                cell.alignment = center_align
                
                if col_a == "Control":
                    cell.fill = control_fill
            
            # Write average
            avg_cell = ws_output.cell(row=output_row_idx, column=3+num_replicates, value=avg_perc)
            avg_cell.border = thin_border
            avg_cell.alignment = center_align
            avg_cell.fill = avg_fill
            
            if col_a == "Control":
                avg_cell.fill = control_fill
            
            output_row_idx += 1
        
        # Adjust column widths
        ws_output.column_dimensions['A'].width = 18
        ws_output.column_dimensions['B'].width = 15
        for col_idx in range(3, 3 + num_replicates + 2):
            ws_output.column_dimensions[get_column_letter(col_idx)].width = 12
    
    # Save output
    wb_output.save(output_path)
    print(f"\nSaved to: {output_path}")
    
    return output_path


if __name__ == "__main__":
    import sys
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Calculate percent inhibition from raw experimental data"
    )
    parser.add_argument(
        "input",
        help="Path to input RawInput.xlsx file"
    )
    parser.add_argument(
        "output",
        nargs="?",
        default="PercInhibition.xlsx",
        help="Path to output file (default: PercInhibition.xlsx)"
    )
    
    args = parser.parse_args()
    
    calculate_percent_inhibition(args.input, args.output)
