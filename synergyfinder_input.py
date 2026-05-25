"""
Generate SynergyFinder Plus Input Format

Converts PercInhibition.xlsx data to SynergyFinder Plus long table format.

Input: PercInhibition.xlsx (from calculate_inhibition.py)
Output: SynergyFinder_input.csv, SynergyFinder_BlockID.csv

Required columns for SynergyFinder Plus:
- block_id: Identifier for drug combination blocks
- drug1: Name of first drug
- drug2: Name of second drug
- conc1: Concentration of first drug
- conc2: Concentration of second drug
- response: % inhibition value
- conc_unit: Unit of concentration (uM)

Each block includes:
- Control (conc1=0, conc2=0, response=0)
- Single drug 1 (conc1>0, conc2=0)
- Single drug 2 (conc1=0, conc2>0)
- Combinations (conc1>0, conc2>0)
"""

import re
from openpyxl import load_workbook
import csv


def parse_drug_conc(value):
    """
    Parse drug_concentration format (e.g., "DrugA_0.78").
    
    Returns:
        tuple: (drug_name, concentration) or (None, None) if not parseable
    """
    if value is None:
        return None, None
    
    value_str = str(value).strip()
    
    # Match pattern: DrugName_Concentration
    match = re.match(r'^(Drug[A-Z])[_\s]+(\d+\.?\d*)$', value_str)
    if match:
        drug_name = match.group(1)
        concentration = float(match.group(2))
        return drug_name, concentration
    
    return None, None


def generate_synergyfinder_input(
    input_path: str,
    output_path: str = "SynergyFinder_input.csv"
):
    """
    Convert PercInhibition.xlsx to SynergyFinder Plus format.
    
    Args:
        input_path: Path to PercInhibition.xlsx
        output_path: Output CSV file path
        
    Output format:
        block_id,drug1,drug2,conc1,conc2,response,conc_unit
    """
    
    # Load input workbook
    wb = load_workbook(input_path)
    
    # Define drug pairs in order
    drug_pairs = [
        ("DrugA", "DrugB"),
        ("DrugA", "DrugC"),
        ("DrugB", "DrugC")
    ]
    
    # Track block summary for BlockID.csv
    block_summary_data = []
    
    # Collect all output rows
    all_output_rows = []
    
    # Track block_id
    block_id = 0
    
    # Store single drug data per experiment
    # {experiment: {drug: [(conc, response), ...]}}
    single_drug_data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Processing: {sheet_name}")
        
        # Read all rows
        rows = list(ws.iter_rows(values_only=True))
        
        # Get header to find Avg_% column index
        header = rows[0]
        avg_col_idx = None
        for i, col in enumerate(header):
            if col == "Avg_%":
                avg_col_idx = i
                break
        
        if avg_col_idx is None:
            print(f"  Warning: Avg_% column not found, skipping sheet")
            continue
        
        # Initialize single drug data for this experiment
        single_drug_data[sheet_name] = {}
        
        # Extract single drug data
        current_drug = None
        for row in rows[1:]:
            col_a = row[0]
            col_b = row[1]
            response = row[avg_col_idx]
            
            # Skip control row
            if col_a == "Control":
                continue
            
            # Check if it's a single drug row
            if col_a is not None and str(col_a).startswith("Drug") and "_" not in str(col_a):
                # Single drug row
                current_drug = str(col_a)
                if current_drug not in single_drug_data[sheet_name]:
                    single_drug_data[sheet_name][current_drug] = []
                
                # Concentration is in col_b
                if col_b is not None and response is not None:
                    single_drug_data[sheet_name][current_drug].append({
                        "conc": col_b,
                        "response": response
                    })
            
            elif col_a is None and current_drug and col_b is not None:
                # Continuation of single drug row
                if response is not None:
                    single_drug_data[sheet_name][current_drug].append({
                        "conc": col_b,
                        "response": response
                    })
            
            # Reset current_drug if we hit a combination row
            if col_a is not None and "_" in str(col_a):
                current_drug = None
        
        # Track block_ids for this experiment
        exp_block_ids = {}
        
        # Process each drug pair for this experiment
        for pair in drug_pairs:
            block_id += 1
            exp_block_ids[pair] = block_id
            
            drug1 = pair[0]
            drug2 = pair[1]
            
            # Add to block summary
            block_summary_data.append({
                "block_id": block_id,
                "drug1": drug1,
                "drug2": drug2,
                "experiment": sheet_name
            })
            
            # 1. Add Control row (conc1=0, conc2=0, response=0)
            all_output_rows.append({
                "block_id": block_id,
                "drug1": drug1,
                "drug2": drug2,
                "conc1": 0,
                "conc2": 0,
                "response": 0,
                "conc_unit": "uM",
                "experiment": sheet_name,
                "row_type": "control"
            })
            
            # 2. Add Drug1 alone rows (conc1>0, conc2=0)
            if drug1 in single_drug_data[sheet_name]:
                for entry in single_drug_data[sheet_name][drug1]:
                    all_output_rows.append({
                        "block_id": block_id,
                        "drug1": drug1,
                        "drug2": drug2,
                        "conc1": entry["conc"],
                        "conc2": 0,
                        "response": entry["response"],
                        "conc_unit": "uM",
                        "experiment": sheet_name,
                        "row_type": "single_drug1"
                    })
            
            # 3. Add Drug2 alone rows (conc1=0, conc2>0)
            if drug2 in single_drug_data[sheet_name]:
                for entry in single_drug_data[sheet_name][drug2]:
                    all_output_rows.append({
                        "block_id": block_id,
                        "drug1": drug1,
                        "drug2": drug2,
                        "conc1": 0,
                        "conc2": entry["conc"],
                        "response": entry["response"],
                        "conc_unit": "uM",
                        "experiment": sheet_name,
                        "row_type": "single_drug2"
                    })
            
            # 4. Add Combination rows (conc1>0, conc2>0)
            # Re-read the worksheet to extract combination data
            current_drug1 = None
            current_conc1 = None
            
            for row in rows[1:]:
                col_a = row[0]
                col_b = row[1]
                response = row[avg_col_idx]
                
                # Skip control row
                if col_a == "Control":
                    continue
                
                # Check if it's a combination row
                if col_a is not None:
                    drug1_check, conc1_check = parse_drug_conc(col_a)
                    
                    if drug1_check and conc1_check:
                        current_drug1 = drug1_check
                        current_conc1 = conc1_check
                        
                        drug2_check, conc2_check = parse_drug_conc(col_b)
                        
                        if drug2_check and conc2_check:
                            # Find matching drug pair
                            pair_key = None
                            for p in drug_pairs:
                                if (p[0] == current_drug1 and p[1] == drug2_check) or \
                                   (p[1] == current_drug1 and p[0] == drug2_check):
                                    pair_key = p
                                    break
                            
                            if pair_key == pair and pair_key in exp_block_ids:
                                all_output_rows.append({
                                    "block_id": exp_block_ids[pair_key],
                                    "drug1": current_drug1,
                                    "drug2": drug2_check,
                                    "conc1": current_conc1,
                                    "conc2": conc2_check,
                                    "response": response,
                                    "conc_unit": "uM",
                                    "experiment": sheet_name,
                                    "row_type": "combination"
                                })
                
                elif col_a is None and col_b is not None and current_drug1:
                    drug2_check, conc2_check = parse_drug_conc(col_b)
                    
                    if drug2_check and conc2_check and current_conc1:
                        pair_key = None
                        for p in drug_pairs:
                            if (p[0] == current_drug1 and p[1] == drug2_check) or \
                               (p[1] == current_drug1 and p[0] == drug2_check):
                                pair_key = p
                                break
                        
                        if pair_key == pair and pair_key in exp_block_ids:
                            all_output_rows.append({
                                "block_id": exp_block_ids[pair_key],
                                "drug1": current_drug1,
                                "drug2": drug2_check,
                                "conc1": current_conc1,
                                "conc2": conc2_check,
                                "response": response,
                                "conc_unit": "uM",
                                "experiment": sheet_name,
                                "row_type": "combination"
                            })
        
        print(f"  Processed {sheet_name}")
    
    # Remove duplicates (keep unique combinations)
    seen = set()
    unique_rows = []
    for row in all_output_rows:
        key = (row["block_id"], row["drug1"], row["drug2"], row["conc1"], row["conc2"])
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)
    
    # Sort rows: by block_id, then by row_type order (control, single, combination)
    row_type_order = {"control": 0, "single_drug1": 1, "single_drug2": 2, "combination": 3}
    unique_rows.sort(key=lambda x: (x["block_id"], row_type_order.get(x["row_type"], 4), x["conc1"], x["conc2"]))
    
    # Write to CSV
    with open(output_path, 'w', newline='') as csvfile:
        fieldnames = ["block_id", "drug1", "drug2", "conc1", "conc2", "response", "conc_unit"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        for row in unique_rows:
            writer.writerow({
                "block_id": row["block_id"],
                "drug1": row["drug1"],
                "drug2": row["drug2"],
                "conc1": row["conc1"],
                "conc2": row["conc2"],
                "response": row["response"],
                "conc_unit": row["conc_unit"]
            })
    
    print(f"\nTotal unique rows: {len(unique_rows)}")
    print(f"Saved to: {output_path}")
    
    # Print summary
    print("\nBlock ID Summary:")
    block_summary = {}
    for row in unique_rows:
        key = (row["block_id"], row["drug1"], row["drug2"], row["experiment"])
        if key not in block_summary:
            block_summary[key] = {"control": 0, "single_drug1": 0, "single_drug2": 0, "combination": 0}
        block_summary[key][row["row_type"]] += 1
    
    for key, counts in sorted(block_summary.items()):
        print(f"  Block {key[0]}: {key[1]}-{key[2]} ({key[3]})")
        print(f"    Control: {counts['control']}, Drug1 alone: {counts['single_drug1']}, Drug2 alone: {counts['single_drug2']}, Combinations: {counts['combination']}")
    
    # Write BlockID.csv
    block_id_path = output_path.replace(".csv", "_BlockID.csv").replace("_input", "")
    with open(block_id_path, 'w', newline='') as csvfile:
        fieldnames = ["block_id", "drug1", "drug2", "experiment"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        for row in block_summary_data:
            writer.writerow({
                "block_id": row["block_id"],
                "drug1": row["drug1"],
                "drug2": row["drug2"],
                "experiment": row["experiment"]
            })
    
    print(f"\nBlock ID summary saved to: {block_id_path}")
    
    return output_path


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else "SynergyFinder_input.csv"
    else:
        input_file = "/Users/kwan/Documents/dengue/DENV_AI/PercInhibition.xlsx"
        output_file = "/Users/kwan/Documents/dengue/DENV_AI/jukjik_v1/SynergyFinder_input.csv"
    
    generate_synergyfinder_input(input_file, output_file)
