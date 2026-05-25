"""
Drug Combination Excel Template Generator

Generates an Excel file with worksheets for drug combination experiments.
Each worksheet contains a template for entering drug combination results.

Input: Excel file with drug concentrations and experiment names
Output: Excel file with all drug pair combinations
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from itertools import combinations


def generate_drug_combination_excel(
    experiments: list,
    output_path: str = "drug_combination_results.xlsx"
):
    """
    Generate Excel file with drug combination template.
    
    Args:
        experiments: List of experiment dictionaries
        output_path: Output Excel file path
    
    Each experiment dict:
        {
            "name": "Experiment Name",      # Worksheet name
            "drug_a": "DrugA",              # First drug name
            "drug_b": "DrugB",              # Second drug name
            "conc_a": [10, 20, 50],         # Concentrations for Drug A
            "conc_b": [5, 10, 20],          # Concentrations for Drug B
            "replicates": 3                 # Number of replicates
        }
    
    Structure per worksheet:
        - Row 1: No Drug control
        - Rows 2-N: Drug A alone (all concentrations)
        - Rows N+1-M: Drug B alone (all concentrations)
        - Remaining rows: All combinations (Drug A x Drug B)
        
        Columns:
        - Col A: Drug A (drug_concentration format)
        - Col B: Drug B (drug_concentration format or "No Drug")
        - Col C onwards: Replicate results (empty for manual entry)
    """
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for exp in experiments:
        ws = wb.create_sheet(title=exp["name"])
        
        drug_a = exp["drug_a"]
        drug_b = exp["drug_b"]
        conc_a = sorted(exp["conc_a"])
        conc_b = sorted(exp["conc_b"])
        replicates = exp["replicates"]
        
        # Style definitions
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        no_drug_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        headers = ["Drug A", "Drug B"] + [f"Rep{i+1}" for i in range(replicates)]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # Track current row
        current_row = 2
        
        # Row 1: No Drug control
        cell_a = ws.cell(row=current_row, column=1, value="No Drug")
        cell_a.fill = no_drug_fill
        cell_a.border = thin_border
        cell_a.alignment = center_align
        
        cell_b = ws.cell(row=current_row, column=2, value="No Drug")
        cell_b.fill = no_drug_fill
        cell_b.border = thin_border
        cell_b.alignment = center_align
        
        # Empty replicate cells
        for col_idx in range(3, 3 + replicates):
            cell = ws.cell(row=current_row, column=col_idx, value="")
            cell.border = thin_border
        
        current_row += 1
        
        # Drug A alone (Drug A concentrations with Drug B = No Drug)
        for conc in conc_a:
            cell_a = ws.cell(row=current_row, column=1, value=f"{drug_a}_{conc}")
            cell_a.border = thin_border
            cell_a.alignment = center_align
            
            cell_b = ws.cell(row=current_row, column=2, value="No Drug")
            cell_b.fill = no_drug_fill
            cell_b.border = thin_border
            cell_b.alignment = center_align
            
            for col_idx in range(3, 3 + replicates):
                cell = ws.cell(row=current_row, column=col_idx, value="")
                cell.border = thin_border
            
            current_row += 1
        
        # Drug B alone (Drug B concentrations in Column A with Drug B column = No Drug)
        for conc in conc_b:
            cell_a = ws.cell(row=current_row, column=1, value=f"{drug_b}_{conc}")
            cell_a.border = thin_border
            cell_a.alignment = center_align
            
            cell_b = ws.cell(row=current_row, column=2, value="No Drug")
            cell_b.fill = no_drug_fill
            cell_b.border = thin_border
            cell_b.alignment = center_align
            
            for col_idx in range(3, 3 + replicates):
                cell = ws.cell(row=current_row, column=col_idx, value="")
                cell.border = thin_border
            
            current_row += 1
        
        # All combinations (Drug A x Drug B)
        for conc_a_val in conc_a:
            for conc_b_val in conc_b:
                cell_a = ws.cell(row=current_row, column=1, value=f"{drug_a}_{conc_a_val}")
                cell_a.border = thin_border
                cell_a.alignment = center_align
                
                cell_b = ws.cell(row=current_row, column=2, value=f"{drug_b}_{conc_b_val}")
                cell_b.border = thin_border
                cell_b.alignment = center_align
                
                for col_idx in range(3, 3 + replicates):
                    cell = ws.cell(row=current_row, column=col_idx, value="")
                    cell.border = thin_border
                
                current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        for col_idx in range(3, 3 + replicates):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")
    return output_path


def read_input_excel(input_path: str):
    """
    Read drug concentrations and experiments from input Excel file.
    
    Args:
        input_path: Path to input Excel file
        
    Returns:
        tuple: (drugs_dict, experiments_list)
            drugs_dict: {drug_name: [conc1, conc2, ...]}
            experiments_list: [{"name": exp_name, "replicates": N}, ...]
    """
    wb = load_workbook(input_path)
    
    # Read DrugAndConcentrations sheet
    drugs = {}
    if "DrugAndConcentrations" in wb.sheetnames:
        ws = wb["DrugAndConcentrations"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            drug_name = row[0]
            if drug_name:
                concentrations = [c for c in row[1:] if c is not None]
                if concentrations:
                    drugs[drug_name] = concentrations
    
    # Read ExperimentName sheet
    experiments = []
    if "ExperimentName" in wb.sheetnames:
        ws = wb["ExperimentName"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            exp_name = row[0]
            replicates = row[1] if row[1] else 3
            if exp_name:
                experiments.append({
                    "name": exp_name,
                    "replicates": int(replicates)
                })
    
    wb.close()
    return drugs, experiments


def generate_from_excel(
    input_path: str,
    output_path: str = "drug_combination_results.xlsx"
):
    """
    Read drug concentrations and experiments from Excel,
    generate output file with all drug pair combinations.
    
    Args:
        input_path: Path to input Excel file (GenExcel.xlsx format)
        output_path: Output Excel file path
        
    Input Excel format:
        Sheet 1: DrugAndConcentrations
            | NAME  | Conc1 | Conc2 | Conc3 | ...
            | DrugA | 0.78  | 1.56  | 3.12  | ...
            | DrugB | 0.09  | 0.19  | 0.39  | ...
            
        Sheet 2: ExperimentName
            | Experiment name | Replicates |
            | Experiment 1    | 4          |
            | Experiment 2    | 3          |
    
    Output:
        One worksheet per (Drug1, Drug2, Experiment) combination
    """
    # Read input
    drugs, experiments = read_input_excel(input_path)
    
    if not drugs:
        raise ValueError("No drugs found in input file")
    if not experiments:
        raise ValueError("No experiments found in input file")
    
    # Generate all unique drug pairs
    drug_names = list(drugs.keys())
    drug_pairs = list(combinations(drug_names, 2))
    
    print(f"Found {len(drugs)} drugs: {list(drugs.keys())}")
    print(f"Found {len(experiments)} experiments: {[e['name'] for e in experiments]}")
    print(f"Generating {len(drug_pairs)} drug pairs: {drug_pairs}")
    
    # Build experiments list for generate_drug_combination_excel
    output_experiments = []
    for drug_a, drug_b in drug_pairs:
        for exp in experiments:
            sheet_name = f"{drug_a}_{drug_b}_{exp['name']}"
            # Truncate sheet name if too long (Excel limit is 31 chars)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            
            output_experiments.append({
                "name": sheet_name,
                "drug_a": drug_a,
                "drug_b": drug_b,
                "conc_a": drugs[drug_a],
                "conc_b": drugs[drug_b],
                "replicates": exp["replicates"]
            })
    
    print(f"Total worksheets to generate: {len(output_experiments)}")
    
    # Generate output
    generate_drug_combination_excel(
        experiments=output_experiments,
        output_path=output_path
    )
    
    return output_path


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else "drug_combination_results.xlsx"
    else:
        input_file = "GenExcel.xlsx"
        output_file = "drug_combination_results.xlsx"
    
    generate_from_excel(input_file, output_file)
